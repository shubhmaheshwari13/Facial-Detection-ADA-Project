import cv2, os
from openpyxl import load_workbook

haar = 'haarcascade_frontalface_default.xml'

n = input('Enter your name: ')
p = input('Enter your password: ')
e = input('Enter your email: ')

while 1:
    m = input('Enter your clearance level: ')
    if m == '1':
        d = 'level_1'
        break
    elif m == '2':
        d = 'level_2'
        break
    else:
        print('Invalid')

path = os.path.join(d, n)
if not os.path.isdir(path):
    os.mkdir(path)

(width, height) = (130, 100)

face_cascade = cv2.CascadeClassifier(haar)
webcam = cv2.VideoCapture(0)

count = 1
while count < 100:
    (_, im) = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 4)
    for (x, y, w, h) in faces:
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))
        cv2.imwrite('% s/% s.png' % (path, count), face_resize)
    count += 1

    cv2.imshow('OpenCV', im)
    key = cv2.waitKey(10)
    if key == 27:
        break

wb = load_workbook('personal_data.xlsx')
ws = wb.active

r = ws.max_row + 1
ws.cell(column=1, row=r, value=n)
ws.cell(column=2, row=r, value=p)
ws.cell(column=3, row=r, value=e)

wb.save(filename='personal_data.xlsx')
wb.close()
